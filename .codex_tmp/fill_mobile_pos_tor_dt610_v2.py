from pathlib import Path
import openpyxl


INPUT = Path(
    "/Users/patrickxu/Library/Containers/com.tencent.WeWorkMac/Data/Documents/Profiles/F832916775EB1B5F9EC73AE338B3F95D/Caches/Files/2026-06/bcf7b1ef17349cf146c46f529477a0f2/New POS_Hardware Requirement V.02 - Mobility.xlsx"
)
OUTPUT_DIR = Path("/Users/patrickxu/product-documents/outputs/mobile_pos_tor")
OUTPUT = OUTPUT_DIR / "New POS_Hardware Requirement V.02 - Mobility_UROVO_DT610_V2_only.xlsx"

CS = "Comply-Standard"
CO = "Comply-Other (Please Explain)"
NC = "Not-Comply"


MODULAR_DT610 = [
    (5, CS, "Proposed equipment is enterprise/commercial class: UROVO DT610 enterprise smartphone, K329 mobile thermal printer, and K180 wireless barcode scanner."),
    (6, CS, "DT610 target industries include Retail; K329 target industries include Retail."),
    (7, CS, "DT610 runs Android 15.0 and supports enterprise applications/App Market."),
    (9, CO, "DT610: Android 15.0, upgradable to Android 19. OS license documentation to be provided separately if required."),
    (10, CS, "DT610: Qualcomm Q-6690 Octa-core 2.0/2.9 GHz processor."),
    (11, CS, "DT610: 8 GB + 128 GB UFS memory."),
    (12, CS, "DT610: 8 GB + 128 GB UFS memory."),
    (13, CS, "DT610: Android 15.0, upgradable to Android 19."),
    (15, CS, "DT610: Wi-Fi 7 (802.11be), 2x2 MU-MIMO, 2.4G/5G/6G support."),
    (16, CS, "DT610: Bluetooth 6.0, BR/EDR + BLE."),
    (17, CS, "DT610: 5G NR Sub-6, Gigabit LTE-A, Hybrid Dual SIM and eSIM support."),
    (18, CO, "DT610 hardware supports Bluetooth 6.0 and USB-C 3.1. External EDC integration to be verified with the client's EDC system."),
    (19, CS, "DT610: 6.0 inch display."),
    (20, CS, "DT610: multi-mode capacitive touch, glove/stylus support and water droplet rejection."),
    (21, CS, "DT610: 1080 x 2160 display resolution."),
    (23, CS, "DT610: sunlight-readable FHD+ display with Corning Gorilla Glass."),
    (24, CS, "K329: thermal line printing mobile printer with Bluetooth 5.0 optional configuration."),
    (25, CS, "K329: Bluetooth 5.0 optional configuration."),
    (26, CS, "K329: print speed up to 120 mm/s."),
    (27, CS, "K329: media width 30-80 mm."),
    (29, CS, "K329: supports receipts and labels."),
    (30, CO, "K329 supports ESC/POS instruction set. Client POS software integration to be verified."),
    (31, CS, "K329: tearing blade for easy tear-off."),
    (41, CS, "K180: dedicated wireless barcode scanner with 1280 x 1080 CMOS image sensor."),
    (42, CS, "K180: Bluetooth 5.2 portable wireless scanner, 95 g."),
    (43, CO, "K180 source confirms 1D/2D decoding. QR/GS1 support is not explicitly stated in the available specification."),
    (44, CO, "K180 source confirms damaged, distorted, laminated, screen-displayed, and low-contrast barcode reading. Reflective barcode reading is not explicitly stated."),
    (45, CS, "K180: physical Scan button."),
    (46, CS, "K180: millisecond decoding with professional scan engine."),
    (48, CS, "K180: buzzer, LED indicators, and vibration motor."),
    (53, CO, "DT610 hardware supports Bluetooth 6.0 and USB-C 3.1. External EDC integration to be verified with the client's EDC system."),
    (56, CS, "DT610 supports Wi-Fi 7 and 5G/Gigabit LTE-A cellular connectivity."),
    (61, CS, "DT610 accessories include 18W wall adapter."),
    (62, CS, "DT610 EMM support includes SOTI, Omnissa Workspace ONE, 42Gears, and safeUEM."),
    (67, CS, "DT610 weight is 222 g."),
    (68, CS, "DT610 supports 1.5 m drop, or 1.8 m with rugged boot; K329 supports 1.5 m drop; K180 supports 1.2 m drop."),
    (88, CS, "DT610 accessories include charging cradles, rugged case, wrist strap, shoulder strap, spare battery and battery charger."),
]


ALL_IN_ONE_I9000S = [
    (91, CS, "i9000S is a rugged smart POS payment terminal with PCI/EMV certification."),
    (92, CS, "i9000S is a smart POS payment terminal for payment processing and retail POS scenarios."),
    (95, CS, "i9000S: Android 13."),
    (96, CS, "i9000S: Octa-core 2.0 GHz processor."),
    (97, NC, "i9000S: RAM is 3 GB; TOR requires at least 4 GB."),
    (98, NC, "i9000S: ROM is 32 GB; TOR requires at least 64 GB."),
    (101, CS, "i9000S: IEEE 802.11 a/b/g/n/ac, 2.4G/5G."),
    (102, CS, "i9000S: Bluetooth 5.0."),
    (103, CS, "i9000S: FDD-LTE/TDD-LTE/WCDMA/GSM; SIM x1 + PSAM x1 + SIM/SAM x1; eSIM option."),
    (104, CO, "i9000S hardware supports Bluetooth 5.0 and Type-C USB. External EDC integration to be verified with the client's EDC system."),
    (105, NC, "i9000S display is 5.0 inch; TOR requires at least 5.5 inch or equivalent."),
    (107, CS, "i9000S: 720 x 1280 display resolution."),
    (110, CS, "i9000S: built-in thermal printer."),
    (111, CS, "i9000S printer: 203 DPI."),
    (112, NC, "i9000S print speed is 50-70 mm/s; TOR requires at least 90 mm/s."),
    (113, CS, "i9000S printer media width: 58 mm."),
    (114, CS, "i9000S has integrated 58 mm receipt printer."),
    (118, CS, "i9000S supports optional 2D scan engine; proposed configuration includes the 2D scan engine."),
    (123, CO, "i9000S hardware supports Bluetooth 5.0 and Type-C USB. External EDC integration to be verified with the client's EDC system."),
    (126, CS, "i9000S supports Wi-Fi and cellular connectivity."),
    (131, NC, "i9000S adapter output is 5V/2A (10W); TOR requires at least 15W charging power."),
    (137, CS, "i9000S weight is 521 g."),
    (138, CS, "i9000S supports 1.5 m drop resistance."),
    (144, CO, "i9000S supports contactless NFC payment and lists PCI, EMV L1/L2, EMV Contactless L1, PayWave, PayPass, and American Express Expresspay. PCI-DSS is not explicitly stated in the available specification."),
]


def fill(ws, row_entries):
    for row, response, remark in row_entries:
        ws.cell(row=row, column=5).value = response
        ws.cell(row=row, column=6).value = remark


OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
wb = openpyxl.load_workbook(INPUT)

for sheet_name in list(wb.sheetnames):
    if sheet_name != "Mobile POS Android V2":
        del wb[sheet_name]

ws = wb["Mobile POS Android V2"]
fill(ws, MODULAR_DT610)
fill(ws, ALL_IN_ONE_I9000S)

wb.save(OUTPUT)
print(OUTPUT)
